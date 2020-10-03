import openpyxl,os
os.chdir('C:\\Users\\Manith Adikari\\Desktop')

workbook = openpyxl.load_workbook('Laptops ranked.xlsx')

print(type(workbook))

print(workbook.get_sheet_names())

sheet1 = workbook.get_sheet_by_name('Sheet1')

print(sheet1['B11'].value)

for i in range(8,12):
    print(i, sheet1.cell(row = i, column =2 ).value)
    

#New workbook
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, Manith'
sheet['A2'] = 'This is a sample excel sheet'
wb.save('Excel_Processing.xlsx')
sheet2 = wb.create_sheet(index=0,title='New sheet1')
sheet2.title = ' Changed sheet2 name'
wb.save()

